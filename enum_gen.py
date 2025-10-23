import pandas as pd
import openpyxl
from openpyxl.cell import Cell
import sys

# ------------------- CONFIGURAÇÕES -------------------
# Altere os valores abaixo de acordo com seu arquivo .xlsx

# 1. Caminho para o arquivo Excel
ARQUIVO_EXCEL = 'CST_cClassTrib.xlsx'

# 2. Nome da aba (sheet) da planilha
NOME_DA_ABA = 'cClassTrib'

# 3. Nome da classe Enum Java de saída
NOME_DO_ENUM = 'CClassTribIBSCBS'

# 4. Nome do arquivo .java de saída
ARQUIVO_DE_SAIDA = f'{NOME_DO_ENUM}.java'

# 5. Mapeamento das colunas:
#    'nome_do_campo_no_java' : 'Nome Exato da Coluna no Excel'
#    A ordem aqui define a ordem dos parâmetros no construtor.
MAPEAMENTO_COLUNAS = {
    'CST_IBS_CBS'           : 'CST-IBS/CBS',
    'cClassTrib'            : 'cClassTrib',
    'percReducaoIBS'        : 'pRedIBS',
    'percReducaoCBS'        : 'pRedCBS',
    'indicadorRedutorBC'    : 'ind_RedutorBC',
    'indTribRegular'        : 'ind_gTribRegular',
    'indCredPresumido'      : 'ind_gCredPresOper',
    'indMonofasico'         : 'ind_gMonoPadrao',
    'indMonoReten'          : 'ind_gMonoReten',
    'indMonoRetido'         : 'ind_gMonoRet',
    'indMonoDiferido'       : 'ind_gMonoDif',
    'indEstornoCred'        : 'ind_gEstornoCred',
    'ind_NFeABI'            : 'indNFeABI',
    'ind_NFe'               : 'indNFe',
    'ind_NFCe'              : 'indNFCe',
    'ind_CTeOS'             : 'indCTeOS',
    'ind_BPe'               : 'indBPe',
    'ind_BPeTA'             : 'indBPeTA',
    'ind_BPeTM'             : 'indBPeTM',
    'ind_NF3e'              : 'indNF3e',
    'ind_NFSe'              : 'indNFSe',
    'ind_NFSe_Via'          : 'indNFSe Via',
    'ind_NFCom'             : 'indNFCom',
    'ind_NFAg'              : 'indNFAg',
    'ind_NFGas'             : 'indNFGas',
}

# 6. Coluna usada para verificar se a linha deve ser ignorada (pelo formato "riscado")
COLUNA_VERIFICACAO_RISCADO = 'cClassTrib'
# ----------------------------------------------------

def formatar_para_string_java(valor):
    if pd.isna(valor) or valor == '':
        return '""'
    # Converte o valor para string, remove espaços extras e o coloca entre aspas
    return f'"{str(valor).strip()}"'

def formatar_para_double_java(valor):
    if pd.isna(valor) or str(valor).strip().upper() == 'N/A' or valor == '':
        return '0.0'
    try:
        return str(float(valor))
    except (ValueError, TypeError):
        return '0.0'

def formatar_para_boolean_java(valor):
    # Valores como '1', 'S', 'TRUE', 'V' (Verdadeiro) são true.
    # Outros (incluindo N/A, '0', 'N', Falso) são false.
    if pd.isna(valor) or str(valor).strip().upper() == 'N/A' or valor == '':
        return 'false'
    
    val_str = str(valor).strip().upper()
    
    if val_str in ['1', '1.0', 'S', 'SIM', 'TRUE', 'V']:
        return 'true'
        
    return 'false'

def obter_dados_sem_riscados():
    """Usa openpyxl para ler o arquivo e filtrar linhas com texto riscado."""
    try:
        workbook = openpyxl.load_workbook(ARQUIVO_EXCEL)
    except FileNotFoundError:
        print(f"Erro: Arquivo '{ARQUIVO_EXCEL}' não encontrado.")
        sys.exit(1)
    except Exception as e:
        print(f"Ocorreu um erro ao abrir o Excel: {e}")
        sys.exit(1)
        
    if NOME_DA_ABA not in workbook.sheetnames:
        print(f"Erro: Aba '{NOME_DA_ABA}' não encontrada no arquivo.")
        print(f"Abas disponíveis: {workbook.sheetnames}")
        sys.exit(1)
        
    sheet = workbook[NOME_DA_ABA]

    all_rows = []
    for row in sheet.iter_rows():
        all_rows.append([cell.value for cell in row])
        
    if not all_rows:
        print(f"Erro: A aba '{NOME_DA_ABA}' está vazia.")
        sys.exit(1)

    header = all_rows[0]
    if COLUNA_VERIFICACAO_RISCADO not in header:
        print(f"Erro: A coluna de verificação '{COLUNA_VERIFICACAO_RISCADO}' não foi encontrada no cabeçalho.")
        print(f"Colunas disponíveis: {header}")
        sys.exit(1)
    
    col_index_to_check = header.index(COLUNA_VERIFICACAO_RISCADO)

    dados_validos = [header] # Começa com o cabeçalho
    
    # Itera pelas linhas de dados (pulando o cabeçalho)
    for row_index, row_cells in enumerate(sheet.iter_rows(min_row=2), start=2):
        cell_to_check: Cell = row_cells[col_index_to_check]
        
        # O atributo 'strike' da fonte nos diz se o texto está riscado
        if not (cell_to_check.font and cell_to_check.font.strike):
            # Se não estiver riscado, adiciona os valores da linha
            dados_validos.append([cell.value for cell in row_cells])

    # Cria e retorna um DataFrame do pandas apenas com os dados válidos
    if len(dados_validos) <= 1:
        print("Aviso: Nenhum dado válido (não riscado) foi encontrado após o cabeçalho.")
        return pd.DataFrame(columns=header)
        
    return pd.DataFrame(dados_validos[1:], columns=dados_validos[0])


def gerar_codigo_enum():
    """Lê a planilha e gera o código Java do Enum."""
    df = obter_dados_sem_riscados()
    if df.empty:
        print("Nenhum dado processado. O arquivo de Enum não será gerado.")
        return
        
    # Converte colunas para string para consistência na formatação
    df = df.astype(str)

    codigo_java = []
    codigo_java.append(f'public enum {NOME_DO_ENUM} {{\n')

    linhas_enum = []
    coluna_cclass_trib = MAPEAMENTO_COLUNAS['cClassTrib']

    for _, row in df.iterrows():
        # Pula linhas onde a coluna principal está vazIA ou é 'nan'
        if pd.isna(row[coluna_cclass_trib]) or str(row[coluna_cclass_trib]).strip().lower() in ['', 'nan', '<na>']:
            continue
            
        nome_constante = f"c{str(row[coluna_cclass_trib]).strip()}"
        
        args = []
        for campo_java, coluna_excel in MAPEAMENTO_COLUNAS.items():
            if coluna_excel not in row:
                print(f"Aviso: Coluna '{coluna_excel}' (para o campo '{campo_java}') não encontrada no DataFrame. Usando valor nulo.")
                valor = None
            else:
                valor = row.get(coluna_excel)
            
            # Determina o tipo de formatação pelo nome do campo Java
            if campo_java.startswith('ind') or campo_java.startswith('indicador'):
                args.append(formatar_para_boolean_java(valor))
            elif campo_java.startswith('perc'):
                args.append(formatar_para_double_java(valor))
            else:
                args.append(formatar_para_string_java(valor))
        
        linhas_enum.append(f"    {nome_constante}({', '.join(args)})")

    if not linhas_enum:
        print("Nenhuma constante de enum foi gerada. Verifique os dados da planilha.")
        codigo_java.append("    // Nenhuma constante gerada.\n")
    else:
        codigo_java.append(',\n'.join(linhas_enum) + ';\n')
    
    # --- Atributos ---
    codigo_java.append("\n    // --- Atributos ---")
    tipos_campos = {}
    for campo_java in MAPEAMENTO_COLUNAS.keys():
        if campo_java.startswith('ind') or campo_java.startswith('indicador'):
            tipo = 'boolean'
        elif campo_java.startswith('perc'):
            tipo = 'double'
        else:
            tipo = 'String'
        tipos_campos[campo_java] = tipo
        codigo_java.append(f"    private final {tipo} {campo_java};")

    # --- Construtor ---
    codigo_java.append('\n    // --- Construtor ---')
    params_construtor = [f"{tipos_campos[campo]} {campo}" for campo in MAPEAMENTO_COLUNAS.keys()]
    codigo_java.append(f"    {NOME_DO_ENUM}({', '.join(params_construtor)}) {{")
    for campo_java in MAPEAMENTO_COLUNAS.keys():
        codigo_java.append(f"        this.{campo_java} = {campo_java};")
    codigo_java.append("    }\n")
    
    # --- Getters ---
    codigo_java.append("    // --- Getters ---")
    for campo_java, tipo in tipos_campos.items():
        nome_getter = f"get{campo_java[0].upper()}{campo_java[1:]}"
        codigo_java.append(f"    public {tipo} {nome_getter}() {{")
        codigo_java.append(f"        return {campo_java};")
        codigo_java.append("    }\n")
        
    codigo_java.append("}")

    # --- Salva o arquivo ---
    try:
        with open(ARQUIVO_DE_SAIDA, 'w', encoding='utf-8') as f:
            f.write('\n'.join(codigo_java))
        print(f"Sucesso! O Enum foi gerado em '{ARQUIVO_DE_SAIDA}' (linhas riscadas foram ignoradas).")
    except Exception as e:
        print(f"Ocorreu um erro ao salvar o arquivo: {e}")

if __name__ == "__main__":
    gerar_codigo_enum()
